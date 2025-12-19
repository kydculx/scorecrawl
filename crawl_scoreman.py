import sys
import subprocess
import os

def check_and_install_packages():
    """필요한 패키지가 설치되어 있는지 확인하고 없으면 설치"""
    required_packages = {
        'pandas': 'pandas',
        'PyQt5': 'PyQt5',
        'playwright': 'playwright',
        'openpyxl': 'openpyxl'
    }
    
    missing_packages = []
    
    for module_name, package_name in required_packages.items():
        try:
            __import__(module_name)
        except ImportError:
            missing_packages.append(package_name)
    
    if missing_packages:
        print("필요한 패키지가 설치되지 않았습니다. 자동으로 설치합니다...")
        print(f"설치할 패키지: {', '.join(missing_packages)}")
        
        # requirements.txt가 있으면 사용, 없으면 개별 설치
        requirements_file = os.path.join(os.path.dirname(__file__), 'requirements.txt')
        if os.path.exists(requirements_file):
            try:
                subprocess.check_call([sys.executable, '-m', 'pip', 'install', '-r', requirements_file])
                print("패키지 설치가 완료되었습니다.")
            except subprocess.CalledProcessError:
                print("패키지 설치 중 오류가 발생했습니다. 수동으로 설치해주세요:")
                print(f"pip install -r requirements.txt")
                sys.exit(1)
        else:
            # requirements.txt가 없으면 개별 설치
            try:
                for package in missing_packages:
                    print(f"{package} 설치 중...")
                    subprocess.check_call([sys.executable, '-m', 'pip', 'install', package])
                print("패키지 설치가 완료되었습니다.")
            except subprocess.CalledProcessError:
                print("패키지 설치 중 오류가 발생했습니다. 수동으로 설치해주세요:")
                print(f"pip install {' '.join(missing_packages)}")
                sys.exit(1)
        
        # playwright의 경우 브라우저도 설치 필요
        try:
            import playwright
            print("Playwright 브라우저 설치 중...")
            subprocess.check_call([sys.executable, '-m', 'playwright', 'install', 'chromium'])
        except:
            pass

# 패키지 확인 및 설치
check_and_install_packages()

import pandas as pd
import re
import json
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QLabel, QLineEdit, QPushButton, 
                             QTextEdit, QMessageBox, QComboBox)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from playwright.sync_api import sync_playwright

def _log(message, log_callback=None):
    """로그 메시지 출력 헬퍼 함수"""
    if log_callback:
        log_callback(message)
    else:
        print(message)

def crawl_scoreman_playwright(url, max_rounds=None, log_callback=None):
    data_list = []

    with sync_playwright() as p:
        # 1. 브라우저 실행 (headless=False로 하면 브라우저 뜨는게 보임 / True는 백그라운드)
        browser = p.chromium.launch(headless=True)
        
        # 2. 새 페이지 생성 및 User-Agent 설정 (차단 방지)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        )
        page = context.new_page()

        _log(f"사이트 접속 중: {url}", log_callback)
        page.goto(url)
        
        # 리그 제목 가져오기
        try:
            league_title_full = page.locator(".info_title").inner_text().strip()
            league_title = league_title_full.split('\n')[0].strip()
            _log(f"리그 제목: {league_title}", log_callback)
        except:
            league_title = "Unknown_League"
            _log("리그 제목을 찾을 수 없어 기본값으로 설정합니다.", log_callback)
        
        # 3. 라운드 정보 가져오기
        # Table2 안에 있는 onclick="changeRound(this)" 속성을 가진 td들을 찾음
        try:
            page.wait_for_selector("#Table2", timeout=5000)
            round_elements = page.locator("#Table2 td[onclick*='changeRound']").all()
            round_texts = [el.inner_text().strip() for el in round_elements if el.inner_text().strip().isdigit()]
            
            # 현재 진행중인 라운드 확인 (class="... round_now ...")
            try:
                current_round_el = page.locator("#Table2 td.round_now").first
                current_round = int(current_round_el.inner_text().strip())
                _log(f"현재 진행중인 라운드: {current_round}", log_callback)
                
                # 라운드 목록에 현재 라운드도 강제로 추가 (onclick이 없어 빠졌을 수 있음)
                round_texts.append(str(current_round))
                
            except:
                _log("현재 라운드 정보를 찾을 수 없습니다. 전체 라운드를 수집합니다.", log_callback)
                current_round = 38 # 못 찾으면 끝까지

            # 1~38 라운드까지 순차적으로 정렬
            round_texts = sorted(list(set(round_texts)), key=int)
            
            # 현재 라운드까지만 필터링
            round_texts = [r for r in round_texts if int(r) <= current_round]
            
            # 테스트용: 수집 라운드 수 제한
            if max_rounds:
                _log(f"테스트 모드: 앞쪽 {max_rounds}개 라운드만 수집합니다.", log_callback)
                round_texts = round_texts[:max_rounds]
            
            _log(f"수집 대상 라운드: {round_texts}", log_callback)
            
            # 배당 회사 목록 가져오기
            company_options = page.locator("#oddsCompany option").all()
            companies = []
            for opt in company_options:
                val = opt.get_attribute("value")
                text = opt.inner_text().strip()
                if val:
                    companies.append((val, text))
            _log(f"수집 대상 배당 회사 ({len(companies)}개): {[n for v, n in companies]}", log_callback)

            # 4. 각 라운드별 데이터 크롤링
            try:
                for round_num in round_texts:
                    _log(f"\n--- {round_num} 라운드 데이터 수집 중 ---", log_callback)
                    
                    try:
                        # 해당 라운드 번호를 가진 버튼(td)을 찾아서 클릭
                        page.locator("#Table2 td").filter(has_text=re.compile(rf"^\s*{round_num}\s*$")).first.click()
                        
                        # 라운드 로딩 대기
                        page.wait_for_timeout(1000)
                        
                        for comp_val, comp_name in companies:
                            try:
                                # 배당 회사 선택
                                page.select_option("#oddsCompany", comp_val)
                                page.wait_for_timeout(50) # 데이터 업데이트 대기

                                # 모든 행(tr) 가져오기
                                all_rows = page.locator("tr").all()
                                
                                round_data_count = 0
                                for idx, row in enumerate(all_rows):
                                    try:
                                        cells = row.locator("td").all()
                                        if len(cells) < 5: continue
                                        
                                        date_time_raw = cells[1].inner_text().strip().split('\n')
                                        date = date_time_raw[0].strip()
                                        time = date_time_raw[1].strip() if len(date_time_raw) > 1 else ''
                                        
                                        home_text = cells[2].inner_text().strip()
                                        away_text = cells[4].inner_text().strip()
                                        
                                        def clean_team_name(text):
                                            """팀 이름에서 HTML 태그, 붙은 숫자, [ ] 사이의 내용 제거"""
                                            # HTML 태그 제거 (inner_text()가 이미 처리하지만 추가 안전장치)
                                            text = re.sub(r'<[^>]+>', '', text)
                                            # [ ] 사이의 모든 내용 제거
                                            text = re.sub(r'\[[^\]]*\]', '', text)
                                            # 팀 이름 앞뒤의 숫자 제거 (예: "1포항", "포항1" -> "포항")
                                            text = re.sub(r'^\d+\s*', '', text)  # 앞의 숫자 제거
                                            text = re.sub(r'\s*\d+$', '', text)  # 뒤의 숫자 제거
                                            return text.strip()
                                        
                                        def parse_team_rank(text):
                                            """팀 이름과 순위 파싱 - [숫자] 형식에서 순위 추출 (숫자만 있을 때만)"""
                                            # HTML 태그 제거
                                            text = re.sub(r'<[^>]+>', '', text)
                                            
                                            # [ ] 사이의 내용 추출
                                            bracket_match = re.search(r'\[([^\]]+)\]', text)
                                            rank = "-"
                                            
                                            if bracket_match:
                                                bracket_content = bracket_match.group(1).strip()
                                                # 순수 숫자만 있는지 확인
                                                if bracket_content.isdigit():
                                                    rank = bracket_content
                                            
                                            # 팀 이름에서 [ ] 제거
                                            cleaned_text = re.sub(r'\[[^\]]*\]', '', text)
                                            # 팀 이름 앞뒤의 숫자 제거
                                            cleaned_text = re.sub(r'^\d+\s*', '', cleaned_text)
                                            cleaned_text = re.sub(r'\s*\d+$', '', cleaned_text)
                                            cleaned_text = cleaned_text.strip()
                                            
                                            return cleaned_text, rank

                                        home_team, home_rank = parse_team_rank(home_text)
                                        away_team, away_rank = parse_team_rank(away_text)
                                        
                                        if home_team.isdigit(): continue

                                        score = cells[3].inner_text().strip()
                                        
                                        # 승/무/패 배당률 (각각 개별 셀에 위치)
                                        win_odds = cells[5].inner_text().strip() if len(cells) > 5 else "-"
                                        draw_odds = cells[6].inner_text().strip() if len(cells) > 6 else "-"
                                        lose_odds = cells[7].inner_text().strip() if len(cells) > 7 else "-"

                                        data_list.append({
                                            "Round": round_num,
                                            "Company": comp_name,
                                            "날짜": date,
                                            "시간": time,
                                            "홈": home_team,
                                            "홈순위": home_rank,
                                            "스코어": score,
                                            "원정": away_team,
                                            "원정순위": away_rank,
                                            "승": win_odds,
                                            "무": draw_odds,
                                            "패": lose_odds
                                        })
                                        round_data_count += 1
                                    except Exception:
                                        continue
                                
                                _log(f"-> [{comp_name}] {round_data_count}개의 데이터 수집 완료", log_callback)
                            except Exception as e_comp:
                                _log(f"  -> {comp_name} 수집 중 에러: {e_comp}", log_callback)
                                continue

                    except Exception as e:
                        _log(f"라운드 {round_num} 처리 중 에러 발생: {e}", log_callback)
                        continue
            except KeyboardInterrupt:
                _log("\n[!] 사용자 중단 요청: 현재까지 수집된 데이터를 저장하고 종료합니다.", log_callback)

        except Exception as e:
            _log(f"라운드 정보를 찾을 수 없습니다: {e}", log_callback)
            browser.close()
            return pd.DataFrame(), ""

        browser.close()

    # 6. DataFrame 반환
    ordered_company_names = [n for v, n in companies]
    return pd.DataFrame(data_list), league_title, ordered_company_names

def post_process_excel(filename):
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment
        
        wb = load_workbook(filename)
        ws = wb.active
        
        # 헤더 스타일 설정 (가운데 정렬)
        align_center = Alignment(horizontal='center', vertical='center')
        
        # 병합할 컬럼 인덱스 (1-based)
        # 1: 순서 (Index), 2: Round, 3: 날짜, 4: 시간
        # 5: 홈, 6: 홈순위, 7: 스코어, 8: 원정, 9: 원정순위
        # 10~: 배당 회사들
        
        merge_cols = range(1, 10) # 1부터 9까지
        
        for col_idx in merge_cols:
            # Index 컬럼(1번) 강제 설정: Pandas가 위치를 A2에 넣거나 누락할 수 있으므로 A1에 '순서' 명시
            if col_idx == 1:
                ws.cell(row=1, column=1).value = "순서"

            ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
            cell = ws.cell(row=1, column=col_idx)
            cell.alignment = align_center

        # 나머지 헤더 셀들도 가운데 정렬
        for row in ws.iter_rows(min_row=1, max_row=2):
            for cell in row:
                cell.alignment = align_center
                
        wb.save(filename)
        print("[서식 적용] 엑셀 헤더 '순서' 및 경기정보 병합 완료")
        
    except ImportError:
        print("[주의] openpyxl이 설치되지 않아 엑셀 후처리를 수행할 수 없습니다.")
    except Exception as e:
        print(f"[주의] 엑셀 후처리 중 오류 발생: {e}")

def process_data(df, title, ordered_companies, log_callback=None):
    """데이터를 처리하고 엑셀 및 JSON 파일로 저장"""
    if df.empty:
        _log("데이터를 가져오지 못했습니다.", log_callback)
        return None
    
    # 파일명으로 사용할 수 없는 문자 제거
    safe_title = re.sub(r'[\\/*?:"<>|]', "", title)
    filename = f"{safe_title.strip()}.xlsx"

    _log("\n[데이터 변환 중...]", log_callback)
    
    # 피벗을 위한 인덱스 컬럼 정의
    index_cols = ["Round", "날짜", "시간", "홈", "홈순위", "스코어", "원정", "원정순위"]
    
    # 중복 제거 (혹시 모를 중복 방지)
    df_unique = df.drop_duplicates(subset=index_cols + ["Company"])
    
    # 피벗 테이블 생성 (행: 경기 정보, 열: 배당 회사, 값: 승/무/패)
    pivot_df = df_unique.pivot(index=index_cols, columns="Company", values=["승", "무", "패"])
    
    # 컬럼 레벨 교환: (승, Bet365) -> (Bet365, 승)
    pivot_df = pivot_df.swaplevel(0, 1, axis=1)
    
    # NOTE: 회사 이름 알파벳 정렬 제거 (사용자 요청: 원래 순서 유지)
    # pivot_df.sort_index(axis=1, level=0, inplace=True)
    
    # 3. 배당 회사 컬럼 순서 정렬 (수집된 순서대로 정렬)
    sorted_cols = []
    
    # 크롤링 단계에서 수집한 순서(ordered_companies)를 사용하여 정렬
    for comp in ordered_companies:
        for type_ in ["승", "무", "패"]:
            if (comp, type_) in pivot_df.columns:
                sorted_cols.append((comp, type_))
    
    pivot_df = pivot_df.reindex(columns=sorted_cols)
    
    # 4. 인덱스 리셋하여 경기 정보를 컬럼으로 변환
    pivot_df.reset_index(inplace=True)
    
    # 라운드, 날짜, 시간 순으로 정렬
    # Round를 숫자로 변환하여 정렬하기 위해 임시 컬럼 생성
    pivot_df['_라운드_정렬'] = pd.to_numeric(pivot_df['Round'], errors='coerce')
    # 날짜를 파싱하여 정렬하기 위해 임시 컬럼 생성
    pivot_df['_날짜_정렬'] = pivot_df['날짜'].apply(lambda x: pd.to_datetime(x, format='%m.%d', errors='coerce') if pd.notna(x) else pd.NaT)
    # Round(숫자), 날짜, 시간 순으로 정렬
    pivot_df = pivot_df.sort_values(by=['_라운드_정렬', '_날짜_정렬', '시간'], na_position='last')
    # 임시 컬럼 삭제
    pivot_df = pivot_df.drop(columns=['_라운드_정렬', '_날짜_정렬'])
    # 인덱스 리셋
    pivot_df.reset_index(drop=True, inplace=True)
    
    # 인덱스 이름 설정
    pivot_df.index.name = "순서"
    
    _log(f"[크롤링 결과] 총 {len(pivot_df)}개의 경기 데이터가 수집되었습니다.", log_callback)
    
    # JSON 저장 및 Excel 생성
    json_filename = f"{safe_title.strip()}.json"
    try:
        # 경기 정보 컬럼 (배당 회사가 아닌 컬럼들)
        match_info_cols = ["Round", "날짜", "시간", "홈", "홈순위", "스코어", "원정", "원정순위"]
        
        # 계층적 JSON 데이터 생성
        def safe_get_value(val):
            """안전하게 값을 추출 (Series 문제 해결)"""
            try:
                if isinstance(val, pd.Series):
                    if len(val) == 0:
                        return None
                    val = val.iloc[0] if len(val) == 1 else val.values[0]
                # None 체크 및 NaN 체크
                if val is None:
                    return None
                # pandas의 NaN 체크 (안전한 방법)
                try:
                    if isinstance(val, float) and (val != val):  # NaN 체크 (NaN != NaN)
                        return None
                except:
                    pass
                return val
            except (ValueError, TypeError, AttributeError):
                return None
        
        json_data = []
        for idx, row in pivot_df.iterrows():
            match_data = {}
            
            # Round 추가
            if "Round" in pivot_df.columns:
                match_data["Round"] = safe_get_value(row["Round"])
            
            # 날짜와 시간을 경기일시로 구조화
            date_val = safe_get_value(row["날짜"]) if "날짜" in pivot_df.columns else None
            time_val = safe_get_value(row["시간"]) if "시간" in pivot_df.columns else None
            
            if date_val is not None or time_val is not None:
                match_data["경기일시"] = {
                    "날짜": date_val,
                    "시간": time_val
                }
            
            # 홈 팀 정보 구조화
            home_team = safe_get_value(row["홈"]) if "홈" in pivot_df.columns else None
            home_rank = safe_get_value(row["홈순위"]) if "홈순위" in pivot_df.columns else None
            
            if home_team is not None or home_rank is not None:
                match_data["홈"] = {
                    "팀": home_team,
                    "순위": home_rank
                }
            
            # 스코어 구조화 (예: "2:1" -> "2-1")
            score = None
            if "스코어" in pivot_df.columns:
                score = safe_get_value(row["스코어"])
            
            if score is not None and str(score).strip():
                score_str = str(score).strip()
                # ":" 또는 "-" 구분자 처리
                if ":" in score_str:
                    score_parts = score_str.split(":")
                    home_score = score_parts[0].strip() if len(score_parts) > 0 else ""
                    away_score = score_parts[1].strip() if len(score_parts) > 1 else ""
                    match_data["스코어"] = f"{home_score}-{away_score}"
                elif "-" in score_str:
                    # 이미 "-" 구분자가 있으면 그대로 사용
                    match_data["스코어"] = score_str
                else:
                    # 구분자가 없으면 그대로 사용
                    match_data["스코어"] = score_str
            else:
                # 스코어가 없거나 빈 값이면 "-" 표시
                match_data["스코어"] = "-"
            
            # 원정 팀 정보 구조화
            away_team = safe_get_value(row["원정"]) if "원정" in pivot_df.columns else None
            away_rank = safe_get_value(row["원정순위"]) if "원정순위" in pivot_df.columns else None
            
            if away_team is not None or away_rank is not None:
                match_data["원정"] = {
                    "팀": away_team,
                    "순위": away_rank
                }
            
            # 배당 회사별로 그룹화하여 계층 구조 생성
            # 경기 정보 컬럼은 제외하고 배당 데이터만 추출
            odds_data = {}
            for col in pivot_df.columns:
                # 컬럼 이름 추출 (문자열이면 그대로, tuple이면 첫 번째 요소)
                if isinstance(col, str):
                    col_name = col
                elif isinstance(col, tuple) and len(col) > 0:
                    col_name = col[0]  # tuple의 첫 번째 요소
                else:
                    col_name = str(col)
                
                # 경기 정보 컬럼은 무조건 제외
                if col_name in match_info_cols:
                    continue
                
                # 배당 데이터만 추출 (tuple 형태: (회사명, "승/무/패"))
                if isinstance(col, tuple) and len(col) == 2:
                    company, odds_type = col[0], col[1]
                    # 회사명도 경기 정보 컬럼이 아닌지 다시 한 번 확인
                    if company not in match_info_cols and odds_type in ["승", "무", "패"]:
                        if company not in odds_data:
                            odds_data[company] = {}
                        value = safe_get_value(row[col])
                        odds_data[company][odds_type] = value
            
            # 배당 데이터가 있으면 추가
            if odds_data:
                match_data["배당"] = odds_data
            
            # 필수 데이터 검증: 홈 팀이나 원정 팀이 있어야 함
            홈 = match_data.get("홈", {})
            원정 = match_data.get("원정", {})
            홈팀 = 홈.get("팀", "") if isinstance(홈, dict) else ""
            원정팀 = 원정.get("팀", "") if isinstance(원정, dict) else ""
            
            # 홈 팀과 원정 팀이 모두 없으면 건너뛰기 (빈 데이터)
            if not 홈팀 and not 원정팀:
                continue
            
            json_data.append(match_data)
        
        # JSON 파일로 저장 (한글 인코딩, 들여쓰기 적용)
        with open(json_filename, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        
        _log(f"JSON 파일 저장 완료: {json_filename}", log_callback)
        
        # JSON 데이터에서 배당 회사 목록 추출 (ordered_companies 순서 유지)
        companies_in_data = set()
        for match in json_data:
            배당 = match.get("배당", {})
            if isinstance(배당, dict):
                companies_in_data.update(배당.keys())
        
        # ordered_companies 순서를 유지하면서 실제 데이터에 있는 회사만 사용
        final_companies = [comp for comp in ordered_companies if comp in companies_in_data]
        # ordered_companies에 없지만 데이터에 있는 회사 추가
        for comp in companies_in_data:
            if comp not in final_companies:
                final_companies.append(comp)
        
        # JSON 데이터를 기반으로 Excel 생성
        create_excel_from_json(json_data, filename, final_companies, log_callback)
        _log(f"엑셀 파일 저장 완료: {filename}", log_callback)
        
    except Exception as e:
        _log(f"[주의] JSON 저장 중 오류 발생: {e}", log_callback)
    
    return filename

def create_excel_from_json(json_data, filename, ordered_companies, log_callback=None):
    """JSON 데이터를 기반으로 Excel 파일 생성"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        
        wb = Workbook()
        ws = wb.active
        
        # 헤더 스타일 설정
        align_center = Alignment(horizontal='center', vertical='center')
        font_bold = Font(bold=True)
        
        # 모든 배당 회사 수집 (ordered_companies 순서 유지)
        all_companies = ordered_companies if ordered_companies else []
        
        # 헤더 생성
        # 첫 번째 행: 메인 헤더
        # 두 번째 행: 서브 헤더
        row1 = []
        row2 = []
        
        # Round
        row1.append("Round")
        row2.append("")
        
        # 경기일시 (날짜, 시간)
        row1.append("경기일시")
        row2.append("날짜")
        row1.append("")
        row2.append("시간")
        
        # 홈 (팀, 순위)
        row1.append("홈")
        row2.append("팀")
        row1.append("")
        row2.append("순위")
        
        # 스코어
        row1.append("스코어")
        row2.append("")
        
        # 원정 (팀, 순위)
        row1.append("원정")
        row2.append("팀")
        row1.append("")
        row2.append("순위")
        
        # 배당 회사들 (각각 승, 무, 패)
        for company in all_companies:
            row1.append(company)
            row2.append("승")
            row1.append("")
            row2.append("무")
            row1.append("")
            row2.append("패")
        
        # 헤더 작성
        ws.append(row1)
        ws.append(row2)
        
        # 헤더 스타일 적용
        for row in ws.iter_rows(min_row=1, max_row=2):
            for cell in row:
                cell.alignment = align_center
                cell.font = font_bold
        
        # 헤더 병합
        # Round
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        
        # 경기일시
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=3)
        
        # 홈
        ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=5)
        
        # 스코어
        ws.merge_cells(start_row=1, start_column=6, end_row=2, end_column=6)
        
        # 원정
        ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=8)
        
        # 배당 회사들
        col_idx = 9
        for company in all_companies:
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+2)
            col_idx += 3
        
        # 헤더 컬럼 수 계산
        expected_cols = 1 + 2 + 2 + 1 + 2 + (len(all_companies) * 3)  # Round + 경기일시(2) + 홈(2) + 스코어 + 원정(2) + 배당(회사수*3)
        
        # 데이터 행 작성
        def safe_value(val):
            """안전하게 값을 변환 (None -> 빈 문자열, 타입 변환)"""
            # None 체크
            if val is None:
                return ""
            
            # NaN 체크 (float 타입인 경우)
            try:
                if isinstance(val, float) and (val != val):  # NaN != NaN
                    return ""
            except:
                pass
            
            # 빈 문자열 체크
            if val == "":
                return ""
            
            # 숫자 타입 변환
            if isinstance(val, (int, float)):
                try:
                    return str(val)
                except:
                    return ""
            
            # 문자열 변환
            try:
                val_str = str(val).strip()
                # 특수 문자나 제어 문자 제거 (엑셀에 문제가 될 수 있는 문자)
                val_str = ''.join(char for char in val_str if ord(char) >= 32 or char in '\n\r\t')
                return val_str
            except:
                return ""
        
        row_count = 0
        for match in json_data:
            # 필수 데이터가 없으면 건너뛰기
            if not isinstance(match, dict):
                continue
            
            # 필수 필드가 있는지 확인 (홈 팀이나 원정 팀이 있어야 함)
            홈 = match.get("홈", {})
            원정 = match.get("원정", {})
            홈팀 = 홈.get("팀", "") if isinstance(홈, dict) else ""
            원정팀 = 원정.get("팀", "") if isinstance(원정, dict) else ""
            
            # 홈 팀과 원정 팀이 모두 없으면 건너뛰기 (빈 데이터)
            if not 홈팀 and not 원정팀:
                continue
            
            row_data = []
            
            # Round
            row_data.append(safe_value(match.get("Round")))
            
            # 경기일시
            경기일시 = match.get("경기일시", {})
            if isinstance(경기일시, dict):
                row_data.append(safe_value(경기일시.get("날짜")))
                row_data.append(safe_value(경기일시.get("시간")))
            else:
                row_data.append("")
                row_data.append("")
            
            # 홈
            홈 = match.get("홈", {})
            if isinstance(홈, dict):
                row_data.append(safe_value(홈.get("팀")))
                row_data.append(safe_value(홈.get("순위")))
            else:
                row_data.append("")
                row_data.append("")
            
            # 스코어
            row_data.append(safe_value(match.get("스코어")))
            
            # 원정
            원정 = match.get("원정", {})
            if isinstance(원정, dict):
                row_data.append(safe_value(원정.get("팀")))
                row_data.append(safe_value(원정.get("순위")))
            else:
                row_data.append("")
                row_data.append("")
            
            # 배당
            배당 = match.get("배당", {})
            if isinstance(배당, dict):
                for company in all_companies:
                    company_odds = 배당.get(company, {})
                    if isinstance(company_odds, dict):
                        row_data.append(safe_value(company_odds.get("승")))
                        row_data.append(safe_value(company_odds.get("무")))
                        row_data.append(safe_value(company_odds.get("패")))
                    else:
                        row_data.append("")
                        row_data.append("")
                        row_data.append("")
            else:
                # 배당이 없으면 모든 회사에 대해 빈 값 추가
                for company in all_companies:
                    row_data.append("")
                    row_data.append("")
                    row_data.append("")
            
            # 컬럼 수 검증
            if len(row_data) != expected_cols:
                _log(f"[경고] 행 {row_count + 1}의 컬럼 수가 맞지 않습니다. 예상: {expected_cols}, 실제: {len(row_data)}", log_callback)
                # 부족한 컬럼은 빈 값으로 채우기
                while len(row_data) < expected_cols:
                    row_data.append("")
                # 초과한 컬럼은 제거
                if len(row_data) > expected_cols:
                    row_data = row_data[:expected_cols]
            
            ws.append(row_data)
            row_count += 1
        
        # 데이터 행 가운데 정렬
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            for cell in row:
                cell.alignment = align_center
        
        wb.save(filename)
        _log("[서식 적용] 엑셀 헤더 병합 및 스타일 적용 완료", log_callback)
        
    except Exception as e:
        _log(f"[주의] Excel 생성 중 오류 발생: {e}", log_callback)
        raise

def post_process_excel(filename, log_callback=None):
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment
        
        wb = load_workbook(filename)
        ws = wb.active
        
        # 헤더 스타일 설정 (가운데 정렬)
        align_center = Alignment(horizontal='center', vertical='center')
        
        # 병합할 컬럼 인덱스 (1-based)
        # 1: 순서 (Index), 2: Round, 3: 날짜, 4: 시간
        # 5: 홈, 6: 홈순위, 7: 스코어, 8: 원정, 9: 원정순위
        # 10~: 배당 회사들
        
        merge_cols = range(1, 10) # 1부터 9까지
        
        for col_idx in merge_cols:
            # Index 컬럼(1번) 강제 설정: Pandas가 위치를 A2에 넣거나 누락할 수 있으므로 A1에 '순서' 명시
            if col_idx == 1:
                ws.cell(row=1, column=1).value = "순서"

            ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
            cell = ws.cell(row=1, column=col_idx)
            cell.alignment = align_center

        # 나머지 헤더 셀들도 가운데 정렬
        for row in ws.iter_rows(min_row=1, max_row=2):
            for cell in row:
                cell.alignment = align_center
                
        wb.save(filename)
        _log("[서식 적용] 엑셀 헤더 '순서' 및 경기정보 병합 완료", log_callback)
        
    except ImportError:
        _log("[주의] openpyxl이 설치되지 않아 엑셀 후처리를 수행할 수 없습니다.", log_callback)
    except Exception as e:
        _log(f"[주의] 엑셀 후처리 중 오류 발생: {e}", log_callback)

class CrawlThread(QThread):
    """크롤링을 실행하는 별도 스레드"""
    log_signal = pyqtSignal(str)
    finished_signal = pyqtSignal(bool, str)  # (성공 여부, 파일명 또는 에러 메시지)
    
    def __init__(self, url, max_rounds=None):
        super().__init__()
        self.url = url
        self.max_rounds = max_rounds
        self.is_running = True
    
    def run(self):
        """크롤링 실행"""
        try:
            self.log_signal.emit(f"크롤링 시작: {self.url}")
            if self.max_rounds:
                self.log_signal.emit(f"최대 라운드 수: {self.max_rounds}")
            
            df, title, ordered_companies = crawl_scoreman_playwright(
                self.url, 
                max_rounds=self.max_rounds,
                log_callback=lambda msg: self.log_signal.emit(msg)
            )
            
            if not self.is_running:
                self.log_signal.emit("크롤링이 중지되었습니다.")
                self.finished_signal.emit(False, "중지됨")
                return
            
            filename = process_data(df, title, ordered_companies, 
                                  log_callback=lambda msg: self.log_signal.emit(msg))
            
            if filename:
                self.log_signal.emit(f"\n완료! 파일이 저장되었습니다: {filename}")
                self.finished_signal.emit(True, filename)
            else:
                self.finished_signal.emit(False, "데이터를 가져오지 못했습니다.")
            
        except Exception as e:
            error_msg = f"크롤링 중 오류가 발생했습니다:\n{str(e)}"
            self.log_signal.emit(f"오류 발생: {str(e)}")
            self.finished_signal.emit(False, error_msg)
    
    def stop(self):
        """크롤링 중지"""
        self.is_running = False
        self.log_signal.emit("크롤링 중지 요청...")

class ScoreCrawlGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("스코어 크롤링 프로그램")
        self.setGeometry(100, 100, 800, 600)
        
        # 중앙 위젯
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 메인 레이아웃
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)
        
        # 도메인 선택
        domain_layout = QHBoxLayout()
        domain_label = QLabel("도메인 선택:")
        self.domain_combo = QComboBox()
        self.domain_combo.addItem("scoreman123.com", "scoreman123.com")
        self.domain_combo.addItem("nowgoal.com", "nowgoal.com")
        self.domain_combo.currentIndexChanged.connect(self.on_domain_changed)
        domain_layout.addWidget(domain_label)
        domain_layout.addWidget(self.domain_combo)
        main_layout.addLayout(domain_layout)
        
        # 리그 선택
        league_layout = QHBoxLayout()
        league_label = QLabel("리그 선택:")
        self.league_combo = QComboBox()
        self.league_combo.setEditable(False)
        self.league_combo.currentIndexChanged.connect(self.on_league_changed)
        
        # 리그 경로 정보 (도메인 제외)
        self.league_paths = [
            ("리그 선택", ""),
            ("K리그 1", "/subleague/15"),
            ("K리그 2", "/subleague/1292"),
            ("호주 A-리그", "/subleague/273"),
            ("프리미어리그", "/league/36"),
            ("세리에 A", "/league/34"),
            ("라리가", "/league/31"),
            ("분데스리가", "/league/8"),
            ("리그 1", "/league/11"),
            ("챔피언스리그", "/cupmatch/103"),
            ("유로파리그", "/cupmatch/113"),
            ("유로파 컨퍼런스리그", "/cupmatch/2187"),
            ("일본 J1리그", "/subleague/25"),
            ("메이저 리그 사커", "/subleague/21"),
            ("CONCACAF챔피언스컵", "/cupmatch/344"),
            ("AFC 아시안컵", "/cupmatch/95"),
        ]
        
        self.update_league_combo()
        
        league_layout.addWidget(league_label)
        league_layout.addWidget(self.league_combo)
        main_layout.addLayout(league_layout)
        
        # URL 입력
        url_layout = QHBoxLayout()
        url_label = QLabel("URL (또는 직접 입력):")
        self.url_input = QLineEdit()
        # 기본 도메인에 맞게 URL 설정
        default_domain = self.domain_combo.currentData()
        self.url_input.setText(f"https://{default_domain}")
        url_layout.addWidget(url_label)
        url_layout.addWidget(self.url_input)
        main_layout.addLayout(url_layout)
        
        # 최대 라운드 수 입력
        round_layout = QHBoxLayout()
        round_label = QLabel("최대 라운드 수 (비워두면 전체):")
        self.max_rounds_input = QLineEdit()
        round_layout.addWidget(round_label)
        round_layout.addWidget(self.max_rounds_input)
        main_layout.addLayout(round_layout)
        
        # 버튼
        button_layout = QHBoxLayout()
        self.start_button = QPushButton("크롤링 시작")
        self.start_button.clicked.connect(self.start_crawl)
        self.stop_button = QPushButton("중지")
        self.stop_button.clicked.connect(self.stop_crawl)
        self.stop_button.setEnabled(False)
        button_layout.addWidget(self.start_button)
        button_layout.addWidget(self.stop_button)
        main_layout.addLayout(button_layout)
        
        # 로그 영역
        log_label = QLabel("진행 상황:")
        main_layout.addWidget(log_label)
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        main_layout.addWidget(self.log_text)
        
        self.crawl_thread = None
    
    def update_league_combo(self):
        """도메인에 따라 리그 콤보박스 업데이트"""
        self.league_combo.clear()
        domain = self.domain_combo.currentData()
        
        for name, path in self.league_paths:
            if path:
                url = f"https://football.{domain}{path}"
            else:
                url = ""
            self.league_combo.addItem(name, url)
    
    def on_domain_changed(self, index):
        """도메인 변경 시 리그 목록 및 URL 업데이트"""
        domain = self.domain_combo.currentData()
        self.update_league_combo()
        # 현재 선택된 리그가 있으면 URL 업데이트
        if self.league_combo.currentIndex() > 0:
            self.on_league_changed(self.league_combo.currentIndex())
        else:
            # 리그가 선택되지 않았으면 기본 도메인 URL로 설정
            self.url_input.setText(f"https://{domain}")
    
    def on_league_changed(self, index):
        """리그 선택 시 URL 업데이트"""
        if index >= 0:
            url = self.league_combo.itemData(index)
            if url:
                self.url_input.setText(url)
    
    def log(self, message):
        """로그 메시지를 텍스트 영역에 추가"""
        if hasattr(self, 'log_text') and self.log_text:
            self.log_text.append(message)
            # 스크롤을 맨 아래로
            scrollbar = self.log_text.verticalScrollBar()
            scrollbar.setValue(scrollbar.maximum())
        else:
            # log_text가 아직 생성되지 않은 경우 출력만
            print(message)
    
    def start_crawl(self):
        """크롤링 시작"""
        if self.crawl_thread and self.crawl_thread.isRunning():
            return
        
        url = self.url_input.text().strip()
        if not url:
            QMessageBox.critical(self, "오류", "URL을 입력해주세요.")
            return
        
        max_rounds = None
        max_rounds_str = self.max_rounds_input.text().strip()
        if max_rounds_str:
            try:
                max_rounds = int(max_rounds_str)
            except ValueError:
                QMessageBox.critical(self, "오류", "최대 라운드 수는 숫자여야 합니다.")
                return
        
        # UI 상태 변경
        self.start_button.setEnabled(False)
        self.stop_button.setEnabled(True)
        self.log_text.clear()
        
        # 크롤링 스레드 시작
        self.crawl_thread = CrawlThread(url, max_rounds)
        self.crawl_thread.log_signal.connect(self.log)
        self.crawl_thread.finished_signal.connect(self.on_crawl_finished)
        self.crawl_thread.start()
    
    def stop_crawl(self):
        """크롤링 중지"""
        if self.crawl_thread and self.crawl_thread.isRunning():
            self.crawl_thread.stop()
    
    def on_crawl_finished(self, success, message):
        """크롤링 완료 처리"""
        self.start_button.setEnabled(True)
        self.stop_button.setEnabled(False)
        
        if success:
            QMessageBox.information(self, "완료", f"크롤링이 완료되었습니다.\n파일: {message}")
        else:
            if message != "중지됨":
                QMessageBox.critical(self, "오류", message)


# --- 실행부 ---
if __name__ == "__main__":
    import sys
    
    # GUI 모드로 실행 (PyQt5)
    app = QApplication(sys.argv)
    window = ScoreCrawlGUI()
    window.show()
    sys.exit(app.exec_())