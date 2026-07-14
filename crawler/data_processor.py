import re
import pandas as pd
from openpyxl.styles import Font, Alignment
from utils.constants import LEAGUE_TABLE


class DataProcessor:
    @staticmethod
    def clean_team_name(text):
        """팀 이름 정제"""
        text = re.sub(r'<[^>]+>', '', text)
        text = re.sub(r'\[[^\]]*\]', '', text)
        text = re.sub(r'^\d+\s*', '', text)
        text = re.sub(r'\s*\d+$', '', text)
        return text.strip()

    @staticmethod
    def parse_team_rank(text):
        """팀 이름과 순위 분리"""
        text = re.sub(r'<[^>]+>', '', text)
        bracket_match = re.search(r'\[([^\]]+)\]', text)
        rank = "-"
        
        if bracket_match:
            bracket_content = bracket_match.group(1).strip()
            if bracket_content.isdigit():
                rank = bracket_content
        
        cleaned_text = re.sub(r'\[[^\]]*\]', '', text)
        cleaned_text = re.sub(r'^\d+\s*', '', cleaned_text)
        cleaned_text = re.sub(r'\s*\d+$', '', cleaned_text)
        return cleaned_text.strip(), rank

    @staticmethod
    def safe_excel_value(val):
        """엑셀 저장용 값 변환"""
        if val is None: return ""
        try:
            if isinstance(val, float) and (val != val): return ""
        except: pass
        if val == "": return ""
        if isinstance(val, (int, float)): return str(val)
        try:
            return ''.join(char for char in str(val).strip() if ord(char) >= 32 or char in '\n\r\t')
        except: return ""

    @classmethod
    def save_results(cls, df, title, log_callback=None, league_name=None, season_name=None):
        if df.empty:
            if log_callback: log_callback("데이터가 없습니다.")
            return None

        safe_title = re.sub(r'[\\/*?:"<>|]', "", title).strip()
        filename = f"{safe_title}.xlsx"

        if log_callback: log_callback("\n[데이터 저장 중...]")

        # 라운드 → 날짜 순 정렬
        df = df.copy()
        df['_라운드_정렬'] = pd.to_numeric(df['Round'], errors='coerce')
        df['_날짜_정렬'] = df['날짜'].apply(
            lambda x: pd.to_datetime(x, format='%m.%d', errors='coerce') if pd.notna(x) else pd.NaT
        )
        df = df.sort_values(by=['_라운드_정렬', '_날짜_정렬', '시간'], na_position='last')
        df.drop(columns=['_라운드_정렬', '_날짜_정렬'], inplace=True)
        df.reset_index(drop=True, inplace=True)
        df.index.name = "순서"

        if log_callback: log_callback(f"[결과] 총 {len(df)}개의 경기 데이터 처리")

        cls._save_excel(df, filename, league_name, season_name)
        if log_callback: log_callback(f"엑셀 저장 완료: {filename}")

        cls._upsert_to_supabase(df, league_name, season_name, log_callback)

        return filename

    @classmethod
    def _upsert_to_supabase(cls, df, league_name, season_name, log_callback):
        if not league_name:
            return

        table_name = LEAGUE_TABLE.get(league_name)
        if not table_name:
            if log_callback:
                log_callback(f"  [Supabase] 리그 '{league_name}'에 매핑된 테이블이 없습니다. 건너뜀")
            return

        try:
            from utils.supabase_client import get_supabase
            supabase = get_supabase()
        except Exception as e:
            if log_callback:
                log_callback(f"  [Supabase] 연결 실패 (건너뜀): {e}")
            return

        rows = []
        for _, row in df.iterrows():
            try:
                r = int(row.get("Round", 0)) if pd.notna(row.get("Round")) else 0
            except (ValueError, TypeError):
                r = 0
            rows.append({
                "시즌": season_name or "",
                "라운드": r,
                "날짜": cls.safe_excel_value(row.get("날짜")),
                "시간": cls.safe_excel_value(row.get("시간")),
                "홈": cls.safe_excel_value(row.get("홈")),
                "홈순위": cls.safe_excel_value(row.get("홈순위")),
                "홈점수": cls.safe_excel_value(row.get("홈스코어")),
                "원정점수": cls.safe_excel_value(row.get("원정스코어")),
                "원정순위": cls.safe_excel_value(row.get("원정순위")),
                "원정": cls.safe_excel_value(row.get("원정")),
                "승": cls.safe_excel_value(row.get("승")),
                "무": cls.safe_excel_value(row.get("무")),
                "패": cls.safe_excel_value(row.get("패")),
                "오버": cls.safe_excel_value(row.get("오버")),
                "오버라인": cls.safe_excel_value(row.get("오버라인")),
                "언더": cls.safe_excel_value(row.get("언더")),
                "핸디캡홈": cls.safe_excel_value(row.get("핸디캡홈")),
                "핸디캡라인": cls.safe_excel_value(row.get("핸디캡라인")),
                "핸디캡원정": cls.safe_excel_value(row.get("핸디캡원정")),
            })

        BATCH_SIZE = 500
        total = len(rows)
        if total == 0:
            if log_callback: log_callback("  [Supabase] 업로드할 데이터가 없습니다.")
            return

        try:
            for i in range(0, total, BATCH_SIZE):
                batch = rows[i:i + BATCH_SIZE]
                supabase.table(table_name).upsert(batch, on_conflict="시즌,날짜,시간,홈,원정").execute()
                if log_callback:
                    log_callback(f"  [Supabase] {table_name} 업로드: {min(i+BATCH_SIZE, total)}/{total}")
            if log_callback:
                log_callback(f"  [Supabase] {table_name} 업로드 완료 ({total}행)")
        except Exception as e:
            if log_callback:
                log_callback(f"  [Supabase] 업로드 중 오류: {e}")

    @classmethod
    def _save_excel(cls, df, filename, league_name=None, season_name=None):
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active

            # 헤더
            header = ["리그", "시즌", "라운드", "날짜", "시간", "홈", "홈순위", "홈점수", "원정점수", "원정순위", "원정",
                       "승", "무", "패",
                       "오버", "오버라인", "언더",
                       "핸디캡홈", "핸디캡라인", "핸디캡원정"]
            ws.append(header)

            header_font = Font(bold=True)
            center_align = Alignment(horizontal='center', vertical='center')
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = center_align

            # 데이터
            col_map = ["Round", "날짜", "시간", "홈", "홈순위", "홈스코어", "원정스코어", "원정순위", "원정",
                       "승", "무", "패",
                       "오버", "오버라인", "언더",
                       "핸디캡홈", "핸디캡라인", "핸디캡원정"]
            for _, row in df.iterrows():
                ws.append([
                    cls.safe_excel_value(league_name),
                    cls.safe_excel_value(season_name),
                    *(cls.safe_excel_value(row[c]) for c in col_map)
                ])

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = center_align

            wb.save(filename)
        except Exception as e:
            raise Exception(f"Excel 생성 실패: {str(e)}")
